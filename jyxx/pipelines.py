# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html
import openpyxl
import re


class JyxxPipeline(object):

    def process_item(self, item, spider):

        if spider.name == 'cggg':
            with open("D:/Document/spider/cggg_row.txt", "r") as f:
                row = f.readline()
                print(row)
            row = int(row) + 1  # 从外来文件读取row以做到逐行写入
            # print('归属地：' + item['level'] + '，公司名：' + item['company'] + '，标题：' + item['title'] + '，项目编号：' + item[
            #     'number'] + '，链接：' + item['href'])
            wb = openpyxl.load_workbook('D:\Document\spider\cggg.xlsx')
            ws = wb.active
            time = re.search('\/201.*\/', item['href']).group(0)[1:-1]
            ws.cell(row=row, column=1, value=time)
            ws.cell(row=row, column=2, value=item['level'])
            ws.cell(row=row, column=3, value=item['company'])
            ws.cell(row=row, column=4, value=item['title'])
            ws.cell(row=row, column=5, value=item['number'])
            ws.cell(row=row, column=6, value=item['href'])
            print(item['href'])
            wb.save('D:\Document\spider\cggg.xlsx')

            with open("D:/Document/spider/cggg_row.txt", "w") as f:
                f.write(str(row))
        if spider.name == 'jggs':
            with open("D:/Document/spider/jggs_row.txt", "r") as f:
                row = f.readline()
                print(row)
            row = int(row) + 1
            # print('归属地：' + item['level'] + '，公司名：' + item['company'] + '，标题：' + item['title'] + '，项目编号：' + item[
            #     'number'] + '，链接：' + item['href'])
            wb = openpyxl.load_workbook('D:\Document\spider\jggs.xlsx')
            ws = wb.active
            time = re.search('\/201.*\/', item['href']).group(0)[1:-1]
            ws.cell(row=row, column=1, value=time)
            ws.cell(row=row, column=2, value=item['level'])
            ws.cell(row=row, column=3, value=item['company'])
            ws.cell(row=row, column=4, value=item['title'])
            ws.cell(row=row, column=5, value=item['number'])
            ws.cell(row=row, column=6, value=item['href'])
            wb.save('D:\Document\spider\jggs.xlsx')

            with open("D:/Document/spider/jggs_row.txt", "w") as f:
                f.write(str(row))
        return item
